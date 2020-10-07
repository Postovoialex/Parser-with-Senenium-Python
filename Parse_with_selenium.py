#Данный парсер реализован с помощью Selenium Web Driver на основе драйвера от Mozila Firefox и был написан только в качестве эксперимента
#Работает данный код не слишком быстро, так как особенностью является загрузка страницы, а только после - фиксация данных, так как все завязано на браузере
# и нет прямого обращения к серверу. Если вы хотите сделать более быстрый парсер используйте модули Beautiful Soup и Requests.
# Данный код будет работать при условии что будет установлен браузер Mozila Firefox 

from selenium import webdriver         
from selenium.webdriver.firefox.options import Options 
from selenium.common.exceptions import NoSuchElementException
import time # Модуль для определения текущего времени   
import openpyxl  # xlsxwriter и openpyxl необходимы для работы с файлами excel
import xlsxwriter 
import datetime   # Модуль для определения даты
from openpyxl.styles import PatternFill # Модуль для окрашивания ячеек в Excel

start_time = time.time() # фиксирую время когда парсер начал работать для того что бы в конце вычислить общее время работы парсера.
now = datetime.datetime.today().strftime("%d.%m.%Y") #Определяю дату в формате день/месяц/год для дальнейшего использования

# прописываю ссылки по которым будет проводится сбор данных. Запись реализована в виде словаря так как полученные данные будут записаны в ексель таблицу
# ключ с словаря определяет название категории и название листа в ексель таблицы.
name_and_url = {  
 'Электролобзики':'https://rozetka.com.ua/jigsaws/c152505/page=1-5;seller=rozetka/'
,'Шуруповерты':'https://rozetka.com.ua/shurupoverty-i-elektrootvertki/c152499/page=1-5;seller=rozetka/'
,'Циркулярные пилы':'https://rozetka.com.ua/pily-i-plitkorezy/c152560/page=1-3;seller=rozetka;23688=12682/'
,'Цепные пилы(эл)':'https://rozetka.com.ua/chainsaws/c155515/page=1-5;seller=rozetka;tip-23970=12927/'
,'Цепные пилы(бенз)':'https://rozetka.com.ua/chainsaws/c155515/page=1-5;seller=rozetka;tip-23970=12928/'
,'Фрезеры':'https://rozetka.com.ua/frezery/c154195/page=1-5;seller=rozetka/'
,'УШМ':'https://rozetka.com.ua/sanders/c152503/page=1-5;seller=rozetka/'
,'Триммеры(бенз)':'https://rozetka.com.ua/trimmers/c155089/page=1-5;seller=rozetka;23888=12862/'
,'Триммеры(эл)':'https://rozetka.com.ua/trimmers/c155089/page=1-5;seller=rozetka;23888=12861/'
,'Сварочные апараты':'https://rozetka.com.ua/svarochnoe-oborudovanie/c152563/page=1-5;seller=rozetka;26517=plazmorez,spotteri,24101/'
,'Сварочные аксы':'https://rozetka.com.ua/svarochnoe-oborudovanie/c152563/page=1-5;seller=rozetka;26517=blok-podachi-provoloki,fiksatori,keysi,sredstvo-dlya-zashchiti-ot-brizg,35585,35590,35725,143147/'
,'Пускозарядные устройства':'https://rozetka.com.ua/chargers/c155293/page=1-5;seller=rozetka/'
,'Перфораторы':'https://rozetka.com.ua/rock_drills/c153621/page=1-5;seller=rozetka/'
,'Отбойные молотки':'https://rozetka.com.ua/jackhammers/c163892/page=1-5;seller=rozetka/'
,'Оприскиватели':'https://rozetka.com.ua/sprayers/c156378/page=1-5;seller=rozetka/'
,'Сварочные маски':'https://rozetka.com.ua/svarochnoe-oborudovanie/c152563/seller=rozetka;26517=24102/'
,'Компрессоры+акс.':'https://rozetka.com.ua/compressors/c162118/page=1-5;seller=rozetka/'
,'Генераторы':'https://rozetka.com.ua/generators/c152564/page=1-5;seller=rozetka/'
,'Паяльники':'https://rozetka.com.ua/payalniki/c4612059/page=1-5;seller=rozetka/'
,'Точильные станки':'https://rozetka.com.ua/grinders/c154219/page=1-5;seller=rozetka/'
,'Пилы и плиткорезы':'https://rozetka.com.ua/pily-i-plitkorezy/c152560/page=1-5;seller=rozetka/'
,'Торцевые пилы':'https://rozetka.com.ua/pily-i-plitkorezy/c152560/page=1-5;seller=rozetka;23688=12683/'
,'Электрорубанки':'https://rozetka.com.ua/elektrorubanki/c152562/page=1-5;seller=rozetka/'
,'Строительные фены':'https://rozetka.com.ua/stroitelnye-feny/c153965/page=1-5;seller=rozetka/'
,'Дрели и миксеры':'https://rozetka.com.ua/drills/c152496/seller=rozetka/'}

workbook = xlsxwriter.Workbook(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) # Создаю файл excel с именем детальный рейтинг + текущая дата что бы имя файлов при ежедневном использовании не повторялись 
for name in name_and_url: 
	worksheet = workbook.add_worksheet("%s" % name) # Создаю листы в ранее созданном Excel файле
	worksheet.set_column('A:A', 5)    # Задаю размер ячейке №
	worksheet.set_column('B:B', 14)	  # Задаю размер ячейке Промо
	worksheet.set_column('C:C', 80)   # Задаю размер ячейке Наименование
	worksheet.set_column('D:D', 12)   # Задаю размер ячейке Старая цена
	worksheet.set_column('E:E', 15)   # Задаю размер ячейке Актуальная цена
	worksheet.set_column('F:F', 17)   # Задаю размер ячейке Цена по промокоду
	worksheet.set_column('G:G', 13)   # Задаю размер ячейке Наличие
	worksheet.set_column('H:H', 13)   # Задаю размер ячейке Отзывы
workbook.close()                      # Сохраняю созданный документ и выхожу с модуля xlsxwriter
print("Файл детальный рейтинг %s создан" % now)

wb = openpyxl.load_workbook(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) # открываю файл детального рейтинга для работы с ним
for name in name_and_url:                                   
	sheet = wb['%s' % name]    #Определяю активный лист для работы с определенным листом               
		
	sheet['A4'] = "№" 				# Вношу в ячейку A4 текст "№" 	
	sheet['B4'] = "Промо"				# Вношу в ячейку В4 текст "Промо" 
	sheet['C4'] = "Наименование"			# Вношу в ячейку С4 текст "Наименование" 		
	sheet['D4'] = "Старая цена"			# Вношу в ячейку D4 текст "Старая цена" 
	sheet['E4'] = "Актуальная цена"			# Вношу в ячейку E4 текст "Актуальная цена" 
	sheet['F4'] = "Цена по промокоду"		# Вношу в ячейку F4 текст "Цена по промокоду" 
	sheet['G4'] = "Наличие"				# Вношу в ячейку G4 текст "Наличие" 
	sheet['H4'] = "Отзывы"				# Вношу в ячейку H4 текст "Отзывы" 

	sheet['A4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке A4 
	sheet['B4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке В4 
	sheet['C4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке С4 
	sheet['D4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке D4 
	sheet['E4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке E4 
	sheet['F4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке F4 
	sheet['G4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке G4 
	sheet['H4'].fill = PatternFill(fgColor="a4ccd0", fill_type = "solid") # задаю фон ячейке H4 

wb.save(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) # сохраняю новосозданный файл
print('Колонки наименованы и закрашены')

#Начинаю работу самого парсера (сбор данных)
options = Options()     #Блокирую запуск браузера
options.headless = True #Если поставить False то откроется браузер Firefox и вы будете наблюдать как поочередно будут открыватся в нем указанные выше страницы 
# Прописываю абсолютный путь к драйверу мозилы 
driver = webdriver.Firefox(options=options, executable_path=r'C:\Users\PC\AppData\Local\Programs\Python\Python37-32\Lib\site-packages\selenium\webdriver\firefox\geckodriver') 
# Можно использовать любое другое местоположение где будет лежать драйвер самого браузера

All_promo = []        #создаю пустой список для наполнения 
All_product_name = []
All_old_price = []
All_now_price = []
All_promocode_price = []
All_avalivable = []
All_feedback = []

wb = openpyxl.load_workbook(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) #открываю созданный файл для того что бы записать его

for name in name_and_url:     # поочередно получаю название ключей с словаря
	link = name_and_url[name] # получаю ссылку с словаря name_and_url с помощью ключа name
	driver.get(link) #Перехожу на полученную сслылку
	
	list_product = driver.find_elements_by_class_name('goods-tile')  #Ищу карточку товара целиком и записываю ее в переменную (получаю данные списком)
	
	for data_list_product in list_product:  # Получаю информацию о промо    
		try:
			actual_promo = data_list_product.find_element_by_class_name('goods-tile__label')  # Ищу промо в полученной ранее товарной позицие 
			All_promo.append (actual_promo.text) 
		except NoSuchElementException: # Ожидаемо будет ошибка так как промо есть не везде
			All_promo.append ('none')
							                                                                                               
		name_product = data_list_product.find_element_by_class_name('goods-tile__title')
		All_product_name.append(name_product.text)
							                                                         
		try:                                       
			old_price = data_list_product.find_element_by_class_name('goods-tile__price_type_old')
			All_old_price.append(old_price.text)
		except NoSuchElementException:
			All_old_price.append('none')
										                                                      
		try:
			actual_price = data_list_product.find_element_by_class_name('goods-tile__price-value')
			All_now_price.append(actual_price.text)
		except NoSuchElementException:
			All_now_price.append('none')
								                                                        
		try:
			price_promo_code = data_list_product.find_element_by_class_name('goods-tile__promo-accent')
			All_promocode_price.append(price_promo_code.text)
		except NoSuchElementException:
			All_promocode_price.append('none')
							                                                             
		actual_available = data_list_product.find_element_by_class_name('goods-tile__availability')
		All_avalivable.append(actual_available.text)
						                                                              
		summ_feedback = data_list_product.find_element_by_class_name('goods-tile__reviews-link')
		All_feedback.append(summ_feedback.text)
	
	print('начинаю наполнение листа %s' % name)

	for x in range(len(list_product)):  #начинаю заполнение листов. len(list_product) определяет количество всех позиций и задает длину цикла
		p = x + 5                       #Цикл считает с нуля, а начинать наполнение ячеек нужно с пятой строчки
		sheet = wb['%s' % name]			#Беру с словаря наименоваие листа и им же определяю активный лист 
		sheet['A%s' % p] = x+1          #Проставляю нумерацию в колонке А числом х, но так как цикл начинаеться с нуля добавляю 1
		sheet['B%s' % p] = All_promo[x]          # Поочередно беру данные с списка All_promo и так же поочередно их заполняю в активный лист
		sheet['C%s' % p] = All_product_name[x]
		sheet['D%s' % p] = All_old_price[x]
		sheet['E%s' % p] = All_now_price[x]
		sheet['F%s' % p] = All_promocode_price[x]
		sheet['G%s' % p] = All_avalivable[x]
		sheet['H%s' % p] = All_feedback[x]

		data = sheet['C%s' % p].value #Беру данные с колонки С 
		data = data.split() # Разделяю полученную строку на отдельные элементы
		for b in data:  # Создаю цикл в котором определяю условие для окрашивания ячеек 
				if b =='Dnipro-M':
					sheet['C%s' % p].fill = PatternFill(fgColor="FD7F00", fill_type = "solid")
				elif b == 'Дніпро-М':
					sheet['C%s' % p].fill = PatternFill(fgColor="FD7F00", fill_type = "solid")
				elif b == 'Foresta':
					sheet['C%s' % p].fill = PatternFill(fgColor="4ea832", fill_type = "solid")

	wb.save(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) # Обязательно! Нужно сохранить документ.

	# Здесь начинается подсчет позиций определенного бренда которые есть в наличии на активном листе в ексель файле
	wb = openpyxl.load_workbook(r'C:\Users\PC\Desktop\Work\Рейтинг\Детальный рейтинг %s.xlsx' % now) #открываю файл
	sheet = wb['%s' % name]  #получаю название листа с словаря выше и делаю его активным
	score_product_in_avalivable = 0 
	for number_colum in range (1,305): # провожу подсчет окрашенных ячеек по указанным ниже условиям
		color_colum = sheet['C%s' % number_colum].fill.start_color.index # определяю индекс цвета ячейки
		avalivable = sheet['G%s' % number_colum].value #получаю данные с ячейки 
		if color_colum == '00FD7F00' and avalivable == 'Есть в наличии' or number_colum == '00fd7f00' and avalivable == 'Есть в наличии': 
			score_product_in_avalivable = score_product_in_avalivable + 1
		elif color_colum =='004ea832' and avalivable == 'Есть в наличии' or color_colum == '004EA832'and avalivable == 'Есть в наличии':
			score_product_in_avalivable = score_product_in_avalivable + 1
	sheet['C1'] = 'В наличии %s' % score_product_in_avalivable

	print(str(score_product_in_avalivable) + name)

	print("лист %s заполнен" % name)
	print()

	All_promo = []          #очищаю списки для их повторного наполнения
	All_product_name = []
	All_old_price = []
	All_now_price = []
	All_promocode_price = []
	All_avalivable = []
	All_feedback = []

driver.quit() # выключаю драйвер 


end_time = time.time()
print('Парсер закончил работу за %s минут' % str((end_time - start_time)/60))
